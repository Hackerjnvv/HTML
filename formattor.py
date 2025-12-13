import os
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter

# Month abbreviations to month numbers
MONTHS = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
}

def clean_name(name):
    """Cleans and normalizes names by removing unnecessary spaces."""
    return ' '.join(name.split())

def extract_data_from_html(html_content):
    """Extracts data from HTML content."""
    soup = BeautifulSoup(html_content, 'lxml')
    data = []

    for card in soup.select('.card'):
        date = student_name = father_name = mother_name = class_info = section = ''

        try:
            date = card.select_one('.date').text.strip() if card.select_one('.date') else ''
            student_name = card.select_one('.card-title').text.strip() if card.select_one('.card-title') else ''

            card_text = card.select_one('.card-text')
            if card_text:
                lines = [line.strip() for line in card_text.stripped_strings]

                if len(lines) >= 1:  # Father's name
                    father_name = clean_name(lines[0].split('/')[0])
                if len(lines) >= 2:  # Mother's name
                    mother_name = clean_name(lines[1])
                if len(lines) >= 3:  # Class info
                    class_info = lines[2].split(':')[-1].split('/')[0].strip()
                if len(lines) >= 4:  # Section
                    section = lines[3].split(':')[-1].strip()

            # Ensure all values are strings for consistent processing
            data.append([
                str(date), str(student_name), str(father_name),
                str(mother_name), str(class_info), str(section)
            ])

        except Exception as e:
            print(f"Error parsing card: {e}")

    return data

def parse_day_month(date_str):
    """Parses a date string in 'DD,MMM' format."""
    try:
        day, month_abbr = date_str.split(',')
        day = int(day.strip())
        month = MONTHS.get(month_abbr.strip(), 0)
        return (month, day)
    except (ValueError, AttributeError):
        return (0, 0) # Return a default tuple if parsing fails

def save_to_excel(data, file_path):
    """Saves the extracted data to an Excel file, avoiding duplicates."""
    try:
        # Create or load the workbook
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Birthday Data"
            sheet.append(["Date", "Student Name", "Father's Name", "Mother's Name", "Class", "Section"])
        
        # Add new data (skip duplicates)
        existing_entries = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Convert all read values to string to match extracted data type
            existing_entries.add(tuple(str(cell) if cell is not None else '' for cell in row))
        
        new_entries_added = 0
        for entry in data:
            if tuple(entry) not in existing_entries:
                sheet.append(entry)
                new_entries_added += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(file_path)
        print(f"Successfully saved {new_entries_added} new entries to Excel: {file_path}")
        return True
    
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return False

def save_to_markdown(data, file_path):
    """Saves the extracted data to a Markdown file as a table, avoiding duplicates."""
    try:
        # Ensure the directory exists
        output_dir = os.path.dirname(file_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        # Read existing entries to avoid duplicates
        existing_entries = set()
        is_new_file = not os.path.exists(file_path)

        if not is_new_file:
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f.readlines()[2:]: # Skip header and separator
                    if not line.strip().startswith('|'):
                        continue
                    # Parse table row: | Val1 | Val2 | -> ['',' Val1 ',' Val2 ','']
                    parts = [p.strip() for p in line.strip().split('|') if p.strip()]
                    if len(parts) == 6:
                        existing_entries.add(tuple(parts))
        
        # Open file in append mode
        with open(file_path, 'a', encoding='utf-8') as f:
            # Write header if it's a new file
            if is_new_file:
                headers = ["Date", "Student Name", "Father's Name", "Mother's Name", "Class", "Section"]
                f.write('| ' + ' | '.join(headers) + ' |\n')
                f.write('|' + '---|' * len(headers) + '\n')
            
            new_entries_added = 0
            for entry in data:
                entry_tuple = tuple(entry)
                if entry_tuple not in existing_entries:
                    # Format entry as a markdown table row
                    md_row = '| ' + ' | '.join(entry_tuple) + ' |\n'
                    f.write(md_row)
                    new_entries_added += 1

        print(f"Successfully saved {new_entries_added} new entries to Markdown: {file_path}")
        return True

    except Exception as e:
        print(f"Error saving to Markdown: {e}")
        return False

def process_html_files(directory):
    """Processes HTML files and saves data to Excel and Markdown."""
    all_data = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.html'):
                file_path = os.path.join(root, file)
                print(f"Processing: {file_path}")
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        all_data.extend(extract_data_from_html(f.read()))
                except Exception as e:
                    print(f"Could not read file {file_path}: {e}")

    # Remove duplicates and sort by date
    # Convert each inner list to a tuple for the set, then back to a list
    unique_data = set(tuple(row) for row in all_data)
    unique_sorted_data = sorted(list(unique_data), key=lambda x: parse_day_month(x[0]))
    
    # Save to Excel file
    excel_path = "Birthday Data Master.xlsx"
    save_to_excel(unique_sorted_data, excel_path)
    
    # Save to Markdown file in the 'html' directory
    markdown_path = os.path.join("html", "Birthday Data Master.md")
    save_to_markdown(unique_sorted_data, markdown_path)

# --- Configuration & Execution ---
# Directory containing HTML files
HTML_DIRECTORY = 'BD'

# Run the process
if __name__ == "__main__":
    if not os.path.isdir(HTML_DIRECTORY):
        print(f"Error: Directory '{HTML_DIRECTORY}' not found. Please create it and place your HTML files inside.")
    else:
        process_html_files(HTML_DIRECTORY)
